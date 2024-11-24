import os
import csv
from openpyxl import load_workbook

def clear_file_contents(filename: str) -> bool:
    """
    Clear the contents of a CSV or Excel file while keeping the headers.

    Args:
        filename (str): The name of the file to clear (CSV or Excel).

    Returns:
        bool: True if the operation was successful, False otherwise.
    """
    try:
        if filename.endswith('.csv'):
            with open(filename, mode="r", newline="") as csvfile:
                reader = csv.reader(csvfile)
                headers = next(reader)  # Read the headers

            with open(filename, mode="w", newline="") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)  # Write the headers back

        elif filename.endswith('.xlsx'):
            workbook = load_workbook(filename)
            sheet = workbook.active

            # Get headers from the first row
            headers = [cell.value for cell in sheet[1]]

            # Clear contents while keeping headers
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None

            # Write headers back to the first row
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)

            workbook.save(filename)

        else:
            print("Unsupported file type. Please provide a CSV or Excel file.")
            return False

        print(f"Contents of {filename} cleared, headers retained.")
        return True

    except Exception as e:
        print(f"Failed to clear contents of {filename}: {e}")
        return False