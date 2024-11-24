import requests
import urllib3
import csv
from datetime import datetime
import os
import clear_file_content as clear  # type: ignore
import tenable.io
import tenable.sc
from openpyxl import load_workbook
import configparser
import sys

urllib3.disable_warnings()

url = "domain/rest/analysis"
headers = {
    "x-apikey": "accesskey=c63e476e7d7a468d83f9e557b1eeb5a1; secretkey=03df3cd43f71419b908cd4a8ca2389ee",
    "Content-Type": "application/json",
}
today = datetime.now().strftime("%Y-%m-%d")
start_offset = 0
batch_size = 100
total_records = 200
csv_file_path = rf"rpt/main_application5.csv"

def clear_file_contents(filename: str) -> bool:
    try:
        if filename.endswith(".csv"):
            with open(filename, mode="r", newline="") as csvfile:
                reader = csv.reader(csvfile)
                headers = next(reader)
            with open(filename, mode="w", newline="") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
        elif filename.endswith(".xlsx"):
            workbook = load_workbook(filename)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)
            workbook.save(filename)
        else:
            print("Unsupported file type. Please provide a CSV or Excel file.")
            return False
        print(f"Contents of {filename} cleared, headers retained.")
        return True
    except Exception as e:
        print(f"Failed to clear contents of {filename}: {e}")
        return False


if not os.path.exists(csv_file_path):
    clear_file_contents(csv_file_path)
    with open(csv_file_path, mode="w", newline="") as file:
        writer = csv.writer(file)
        headers = [
            "pluginID",
            "vulnUUID",
            "severityid",
            "severity_name",
            "severity_description",
            "pluginName",
            "protocol",
            "IP Address",
            "dnsName",
            "netbiosName",
            "plugin_output",
            "cve",
            "firstSeen",
            "lastSeen",
            "description",
            "hasBeenMitigated",
            "Cross References",
            "repository id",
            "repository name",
        ]
        writer.writerow(headers)

# Get the timestamp from command line arguments
if len(sys.argv) != 2:
    print("Usage: python tanable.py <timestamp>")
    sys.exit(1)

last_seen_epoch = sys.argv[1]
endtime = datetime.now().timestamp().hex()    
timestamp = f"  {last_seen_epoch}   {endtime}"

while start_offset < total_records:
    end_offset = start_offset + batch_size - 1
    payload = {
        "query": {
            "type": "vuln",
            "tool": "vulndetails",
            "sourceType": "cumulative",
            "startOffset": start_offset,
            "endOffset": end_offset,
            "filters": [
                {
                    "filterName": "lastSeen",
                    "operator": "=",
                    "value": timestamp,  # Use the timestamp from command line
                },
                {
                    "filterName": "pluginID",
                    "operator": "=",
                    "value": "0-999999",
                },
                {"filterName": "severity", "operator": "!=", "value": "0"},
            ],
        },
        "sortDir": "asc",
        "sortField": "lastSeen",
        "sourceType": "cumulative",
        "type": "vuln",
    }

    response = requests.post(url, headers=headers, json=payload, verify=False)

    if response.status_code == 200:
        data = response.json()
        totalRecords = data.get("response", {}).get("totalRecords")
        total_records = int(totalRecords)
        print("Total Records: ", totalRecords)
        results = data.get("response", {}).get("results", [])

        with open(csv_file_path, mode="a", newline="") as file:
            writer = csv.writer(file)
            for item in results:
                row = [
                    item.get("pluginID"),
                    item.get("vulnUUID"),
                    item.get("severity", {}).get("id"),
                    item.get("severity", {}).get("name"),
                    item.get("severity", {}).get("description"),
                    item.get("pluginName"),
                    item.get("protocol"),
                    item.get("ip"),
                    item.get("dnsName"),
                    item.get("netbiosName"),
                    item.get("pluginText"),
                    item.get("cve"),
                    item.get("firstSeen"),
                    item.get("lastSeen"),
                    item.get("description"),
                    item.get("hasBeenMitigated"),
                    item.get("xref"),
                    item.get("repository", {}).get("id"),
                    item.get("repository", {}).get("name"),
                ]
                writer.writerow(row)

        print(f"Batch {start_offset}-{end_offset} written to {csv_file_path}")
    else:
        print("Error:", response.status_code, response.text)
        break

    start_offset += batch_size
